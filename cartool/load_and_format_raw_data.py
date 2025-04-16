import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries


def unmerge_and_fill_cells(path_in, path_out):

    wb = load_workbook(path_in)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        merge_list = []
        for merge in sheet.merged_cells.ranges:
            merge_list.append(merge)

        for group in merge_list:
            print(f"Changing merged cells {group.coord} in sheet {sheet_name}")
            min_col, min_row, max_col, max_row = range_boundaries(group.coord)
            top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
            sheet.unmerge_cells(str(group))
            for row in sheet.iter_rows(
                min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row
            ):
                for cell in row:
                    cell.value = top_left_cell_value

    wb.save(path_out)


def main():
    """
    Main function to load and format the data
    """
    path = "data/interim/Main_Results_CARTool_2025-04-15.xlsx"
    path_out = "data/processed/Main_Results_CARTool_2025-04-15.xlsx"
    unmerge_and_fill_cells(path_in=path, path_out=path_out)


if __name__ == "__main__":
    main()
